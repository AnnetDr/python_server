from flask import Flask, render_template, url_for, request
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
import pandas as pd
import matplotlib.pyplot as plt


app =  Flask(__name__)

ALLOWED_EXTENSIONS = {'xls'}
EXCEL_FILES_UPLOADED_PATH = "data//excel//"

        
def add_uploaded_file_to_list(full_path):
    database = open('database.txt','a+')
    file_name = os.path.basename(full_path)
    database.seek(0)
    saved_names = database.read().splitlines()
    name = ""
    for name in saved_names:
        if name == file_name:
            file_name = ""
    if file_name != "":
        database.write(file_name)
        database.write("\n")
    database.close()
        

def get_upload_file_list():
    database = open('database.txt','a+')
    database.seek(0)
    file_list = []
    saved_names = database.read().splitlines()
    for name in saved_names:
        file_list.append(name)
    database.close()
    return file_list
     

def read_excel_file(file_name):
    work_book = load_workbook(filename=file_name, data_only=True)
    name_list = work_book.get_sheet_names();
    sheet = work_book.worksheets[0]
    sheet_name = sheet.title
    list_rows = []
    for row in sheet.iter_rows(min_row=0):
        one_row = []
        for cell in row:
            if cell.value is None:
                cell.value = "-"
            one_row.append(cell.value)
        list_rows.append(one_row)
    return list_rows

def draw_graph(excel_file_name): 
    exel_data = pd.read_excel(excel_file_name)    
    fig = plt.figure()
    plt.plot(range(2))
    plt.plot(exel_data['ШБУ'])
    plt.title('Значения ШБУ')
    fig.savefig('static/graph.png', dpi=fig.dpi)


@app.route('/') 
def initialize():
    if not os.path.exists(EXCEL_FILES_UPLOADED_PATH):
        full_data_path = os.path.abspath(os.getcwd()) + "//" + EXCEL_FILES_UPLOADED_PATH
        os.makedirs(full_data_path)
    return render_template('components.html', error_message="Выберите файл", uploads=get_upload_file_list())


@app.route('/', methods=['POST']) 
def initialize_asdf1():
    global new_list
    if request.method == 'POST':    
        try:   
            if "excel_file" in request.files:
                exel_file = request.files.get('excel_file')
                file_name = secure_filename(exel_file.filename)
                uploaded_file_full_name = EXCEL_FILES_UPLOADED_PATH + file_name
                exel_file.save(uploaded_file_full_name)
                add_uploaded_file_to_list(uploaded_file_full_name)
                draw_graph(uploaded_file_full_name)

                return render_template('components.html',
                                       excel_cells = read_excel_file(uploaded_file_full_name),
                                       error_message = file_name,
                                       uploads = get_upload_file_list(),
                                       graph_img = True)
            
            else:
                graph_name = request.form['graph_name']
                if graph_name != "":
                    full_graph_name = EXCEL_FILES_UPLOADED_PATH + graph_name
                    draw_graph(full_graph_name)
                    return render_template('components.html', 
                                        error_message = graph_name, 
                                        excel_cells = read_excel_file(full_graph_name),
                                        uploads = get_upload_file_list(),
                                        graph_img = True)
                
        except:
            error_message="Выберите файл"
            return render_template('components.html', error_message=error_message)
            
    return render_template('components.html', uploads=get_upload_file_list())
    
